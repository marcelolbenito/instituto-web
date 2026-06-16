#!/usr/bin/env python3
"""
Activa alumnos de postítulo/postgrado desde Excel (solo activación + concepto).

Excel esperado (hoja 1, encabezados fila 3):
  A = articulo_id (concepto a asignar en alumno_articulo)
  B = apellido
  C = nombre
  D = DNI

No desactiva a nadie, no genera cuotas ni pagos.

Uso (producción):
  1. Backup de BD
  2. python tools/activar_postitulo_excel.py --excel "LISTADO COMPLETO POSTITULO.xlsx" --dry-run
  3. python tools/activar_postitulo_excel.py --excel "LISTADO COMPLETO POSTITULO.xlsx"

Variables de entorno:
  MYSQL_HOST (si está definido: mariadb directo por SSH; si no: docker exec)
  MYSQL_DOCKER_CONTAINER, MYSQL_USER, MYSQL_PASSWORD, MYSQL_DATABASE
"""
from __future__ import annotations

import argparse
import os
import subprocess
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instalar: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = ROOT / "LISTADO COMPLETO POSTITULO.xlsx"


def load_env_file() -> None:
    env_path = ROOT / ".env"
    if not env_path.is_file():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def mysql_exec(sql: str, *, fetch: bool = False) -> str:
    user = os.environ.get("MYSQL_USER", "instituto")
    password = os.environ.get("MYSQL_PASSWORD", "instituto")
    db = os.environ.get("MYSQL_DATABASE", "instituto")
    host = os.environ.get("MYSQL_HOST", "").strip()

    if host:
        cmd = [
            "mariadb",
            f"-h{host}",
            f"-u{user}",
            f"-p{password}",
            db,
            "-N",
            "-e",
            sql,
        ]
    else:
        container = os.environ.get("MYSQL_DOCKER_CONTAINER", "instituto-db")
        cmd = [
            "docker",
            "exec",
            container,
            "mariadb",
            f"-u{user}",
            f"-p{password}",
            db,
            "-N",
            "-e",
            sql,
        ]
    try:
        out = subprocess.check_output(cmd, stderr=subprocess.STDOUT, text=True, errors="replace")
    except subprocess.CalledProcessError as e:
        raise RuntimeError(e.output or str(e)) from e
    return out.strip() if fetch else ""


def sql_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace("'", "''")


def column_exists(table: str, column: str) -> bool:
    t = sql_escape(table)
    c = sql_escape(column)
    out = mysql_exec(
        f"""
        SELECT COUNT(*) FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = '{t}'
          AND COLUMN_NAME = '{c}';
        """,
        fetch=True,
    )
    return out.strip() == "1"


def read_excel(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict] = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i < 3 or not r:
            continue
        dni_raw = r[3] if len(r) > 3 else None
        if dni_raw is None:
            continue
        art_raw = r[0]
        if art_raw is None:
            continue
        try:
            articulo_id = int(art_raw)
        except (TypeError, ValueError):
            continue
        if articulo_id <= 0:
            continue
        dni = str(int(dni_raw)) if isinstance(dni_raw, (int, float)) else str(dni_raw).strip()
        if not dni:
            continue
        rows.append(
            {
                "articulo_id": articulo_id,
                "dni": dni,
                "apellido": str(r[1] or "").strip() if len(r) > 1 else "",
                "nombre": str(r[2] or "").strip() if len(r) > 2 else "",
            }
        )
    return rows


def lookup_rows(rows: list[dict]) -> list[dict]:
    """Enriquece cada fila con alumno_id, estado actual y validez del artículo."""
    out: list[dict] = []
    for row in rows:
        dni = sql_escape(row["dni"])
        art = int(row["articulo_id"])
        alumno_line = mysql_exec(
            f"""
            SELECT id, activo, COALESCE(nombre_completo, ''), COALESCE(tipo_alumno, 'regular')
            FROM alumnos
            WHERE TRIM(documento) = TRIM('{dni}')
            LIMIT 1;
            """,
            fetch=True,
        )
        art_line = mysql_exec(
            f"""
            SELECT id, COALESCE(detalle, ''), activo
            FROM articulos
            WHERE id = {art}
            LIMIT 1;
            """,
            fetch=True,
        )
        aa_line = ""
        if alumno_line:
            parts = alumno_line.split("\t")
            alumno_id = int(parts[0])
            aa_line = mysql_exec(
                f"""
                SELECT 1 FROM alumno_articulo
                WHERE alumno_id = {alumno_id} AND articulo_id = {art}
                LIMIT 1;
                """,
                fetch=True,
            )

        enriched = dict(row)
        if not alumno_line:
            enriched["error"] = "DNI no encontrado en alumnos"
            enriched["alumno_id"] = None
        else:
            parts = alumno_line.split("\t")
            enriched["alumno_id"] = int(parts[0])
            enriched["activo_antes"] = int(parts[1])
            enriched["nombre_bd"] = parts[2]
            enriched["tipo_antes"] = parts[3]
            if not art_line:
                enriched["error"] = f"articulo_id {art} no existe"
            elif art_line.split("\t")[2] != "1":
                enriched["error"] = f"articulo_id {art} inactivo en articulos"
            else:
                enriched["articulo_detalle"] = art_line.split("\t")[1]
                enriched["ya_tiene_concepto"] = aa_line == "1"
        out.append(enriched)
    return out


def print_report(enriched: list[dict], *, dry_run: bool) -> int:
    ok = [r for r in enriched if not r.get("error")]
    err = [r for r in enriched if r.get("error")]
    activar = [r for r in ok if int(r.get("activo_antes", 0)) != 1]
    ya_activos = [r for r in ok if int(r.get("activo_antes", 0)) == 1]
    nuevos_concepto = [r for r in ok if not r.get("ya_tiene_concepto")]

    print(f"\n=== {'DRY-RUN — ' if dry_run else ''}Resumen ===")
    print(f"Filas Excel:        {len(enriched)}")
    print(f"OK (matchean BD):   {len(ok)}")
    print(f"Errores:            {len(err)}")
    print(f"A activar (inactivos): {len(activar)}")
    print(f"Ya activos:         {len(ya_activos)}")
    print(f"Concepto a vincular: {len(nuevos_concepto)} filas (INSERT IGNORE alumno_articulo)")

    by_art: dict[int, int] = {}
    for r in ok:
        aid = int(r["articulo_id"])
        by_art[aid] = by_art.get(aid, 0) + 1
    if by_art:
        print("Por articulo_id:", ", ".join(f"{k}={v}" for k, v in sorted(by_art.items())))

    if err:
        print("\nErrores (hasta 25):")
        for r in err[:25]:
            nom = f"{r.get('apellido', '')}, {r.get('nombre', '')}".strip(", ")
            print(f"  DNI {r['dni']} | art {r['articulo_id']} | {r['error']} | {nom}")

    if dry_run and activar[:10]:
        print("\nMuestra a activar (hasta 10):")
        for r in activar[:10]:
            print(
                f"  id={r['alumno_id']} DNI {r['dni']} art={r['articulo_id']} "
                f"| {r.get('nombre_bd', '')}"
            )

    return len(err)


def aplicar(enriched: list[dict]) -> None:
    has_tipo = column_exists("alumnos", "tipo_alumno")
    has_estado = column_exists("alumnos", "estado_cuenta")
    has_fecha_ina = column_exists("alumnos", "fecha_inactivacion")

    tipo_sql = ", a.tipo_alumno = 'postgrado'" if has_tipo else ""
    estado_sql = ", a.estado_cuenta = 'activo'" if has_estado else ""
    fecha_sql = ", a.fecha_inactivacion = NULL" if has_fecha_ina else ""

    activados = 0
    conceptos = 0
    for row in enriched:
        if row.get("error") or row.get("alumno_id") is None:
            continue
        alumno_id = int(row["alumno_id"])
        articulo_id = int(row["articulo_id"])
        dni = sql_escape(row["dni"])

        mysql_exec(
            f"""
            UPDATE alumnos a
            SET a.activo = 1
            {estado_sql}
            {fecha_sql}
            {tipo_sql}
            WHERE a.id = {alumno_id}
              AND TRIM(a.documento) = TRIM('{dni}');
            """
        )
        if int(row.get("activo_antes", 0)) != 1:
            activados += 1

        mysql_exec(
            f"""
            INSERT IGNORE INTO alumno_articulo (alumno_id, articulo_id)
            VALUES ({alumno_id}, {articulo_id});
            """
        )
        if not row.get("ya_tiene_concepto"):
            conceptos += 1

    print(f"\nAplicado: {activados} alumnos pasaron a activo; {conceptos} vínculos alumno_articulo nuevos.")


def main() -> None:
    load_env_file()
    parser = argparse.ArgumentParser(description="Activar postgrado/postítulo desde Excel")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Ruta al .xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Solo informe, sin escribir en BD")
    parser.add_argument(
        "--yes",
        action="store_true",
        help="Aplicar filas OK aunque haya errores (sin preguntar; útil por SSH)",
    )
    args = parser.parse_args()

    if not args.excel.is_file():
        print(f"No existe el archivo: {args.excel}")
        sys.exit(1)

    rows = read_excel(args.excel)
    print(f"Filas leídas del Excel: {len(rows)}")
    if not rows:
        print("Sin filas válidas (revisar columnas A=articulo_id, D=DNI, encabezados fila 3).")
        sys.exit(1)

    enriched = lookup_rows(rows)
    n_err = print_report(enriched, dry_run=args.dry_run)

    if args.dry_run:
        print("\nEjecutá sin --dry-run para aplicar.")
        if n_err:
            sys.exit(2)
        return

    if n_err and not args.yes:
        print("\nHay errores. Corregí el Excel o la BD antes de aplicar, o revisá la lista arriba.")
        print("Para aplicar las filas OK sin preguntar: agregá --yes")
        resp = input("¿Aplicar igual para las filas OK? [s/N]: ").strip().lower()
        if resp not in ("s", "si", "sí", "y", "yes"):
            sys.exit(2)

    aplicar(enriched)
    print("Listo. Verificá en Alumnos (filtro Activos) y Conceptos por alumno si hace falta.")


if __name__ == "__main__":
    main()
