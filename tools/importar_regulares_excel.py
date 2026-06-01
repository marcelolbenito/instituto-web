#!/usr/bin/env python3
"""
Importa alumnos regulares desde Excel (reemplazo operativo post-purge Fox).

Flujo recomendado (dev o producción):
  1. Backup de BD
  2. sql/migracion/23_purge_pagos_legacy_fox.sql
  3. python tools/importar_regulares_excel.py --excel "ALUMNOS REGULARES (MARCELO) (1).xlsx"
  4. php tools/recalcular_saldos_cli.php

Variables de entorno opcionales:
  MYSQL_DOCKER_CONTAINER=instituto-db  (default)
  MYSQL_USER, MYSQL_PASSWORD, MYSQL_DATABASE
  OPERATIVO_ANIO_DESDE=2026
"""
from __future__ import annotations

import argparse
import os
import re
import subprocess
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instalar: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = ROOT / "ALUMNOS REGULARES (MARCELO) (1).xlsx"

MES_ESPECIAL = frozenset({"BECA", "MATRICULA", "MATRÍCULA"})

# Excel con typo → DNI en alumnos (revisar con administración antes de producción).
DNI_CORRECCIONES: dict[str, str] = {
    "42634242": "42634424",  # DUARTE, LOURDES MICAELA
    "46155010": "40155010",  # JOJOT, FRANCO
    "37588607": "37588637",  # BARBIERI, CARLA
    "44924357": "44924375",  # BLASICH, SEBASTIAN
    "40211309": "40211038",  # MARTINEZ, MARISEL
    "45899843": "46899843",  # SOSA, PATRICIO
}


def load_env_file() -> None:
    env_path = ROOT / ".env"
    if not env_path.is_file():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def mysql_exec(sql: str, *, fetch: bool = False) -> str:
    container = os.environ.get("MYSQL_DOCKER_CONTAINER", "instituto-db")
    user = os.environ.get("MYSQL_USER", "instituto")
    password = os.environ.get("MYSQL_PASSWORD", "instituto")
    db = os.environ.get("MYSQL_DATABASE", "instituto")

    cmd = [
        "docker",
        "exec",
        container,
        "mariadb",
        f"-u{user}",
        f"-p{password}",
        db,
        "-N",
        "-e",
        sql,
    ]
    try:
        out = subprocess.check_output(cmd, stderr=subprocess.STDOUT, text=True, errors="replace")
    except subprocess.CalledProcessError as e:
        raise RuntimeError(e.output or str(e)) from e
    return out.strip() if fetch else ""


def sql_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace("'", "''")


def parse_mes_abonado(raw) -> tuple[int | None, int | None, str]:
    if raw is None:
        return None, None, ""
    if isinstance(raw, datetime):
        return raw.year, raw.month, raw.strftime("%Y-%m")
    text = str(raw).strip().upper()
    if text in MES_ESPECIAL:
        return None, None, text
    m = re.match(r"^(\d{4})-(\d{1,2})$", text)
    if m:
        return int(m.group(1)), int(m.group(2)), text
    return None, None, text


def read_excel(path: Path, *, corregir_dni: bool = False) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict] = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i < 3 or not r or not isinstance(r[0], (int, float)):
            continue
        dni_raw = r[3]
        if dni_raw is None:
            continue
        dni_excel = str(int(dni_raw)) if isinstance(dni_raw, (int, float)) else str(dni_raw).strip()
        dni = DNI_CORRECCIONES[dni_excel] if corregir_dni and dni_excel in DNI_CORRECCIONES else dni_excel
        anio, mes, mes_raw = parse_mes_abonado(r[5])
        rows.append(
            {
                "dni": dni,
                "dni_excel": dni_excel,
                "apellido": str(r[1] or "").strip(),
                "nombre": str(r[2] or "").strip(),
                "carrera": str(r[4] or "").strip(),
                "mes_raw": mes_raw,
                "mes_anio": anio,
                "mes_mes": mes,
            }
        )
    return rows


def ensure_staging_table() -> None:
    sql_path = ROOT / "sql/migracion/24_staging_excel_regulares.sql"
    mysql_exec(sql_path.read_text(encoding="utf-8"))


def load_staging(rows: list[dict], *, truncate: bool) -> None:
    if truncate:
        mysql_exec("TRUNCATE TABLE staging_excel_regulares;")
    batch = []
    for row in rows:
        anio_sql = "NULL" if row["mes_anio"] is None else str(row["mes_anio"])
        mes_sql = "NULL" if row["mes_mes"] is None else str(row["mes_mes"])
        batch.append(
            "INSERT INTO staging_excel_regulares "
            "(dni, apellido, nombre, carrera, mes_abonado_raw, mes_abonado_anio, mes_abonado_mes) VALUES ("
            f"'{sql_escape(row['dni'])}', "
            f"'{sql_escape(row['apellido'])}', "
            f"'{sql_escape(row['nombre'])}', "
            f"'{sql_escape(row['carrera'])}', "
            f"'{sql_escape(row['mes_raw'])}', "
            f"{anio_sql}, {mes_sql});"
        )
    chunk = 50
    for i in range(0, len(batch), chunk):
        mysql_exec("\n".join(batch[i : i + chunk]))


def match_alumnos() -> None:
    mysql_exec(
        """
        UPDATE staging_excel_regulares s
        LEFT JOIN alumnos a ON TRIM(a.documento) = TRIM(s.dni)
        SET s.alumno_id = a.id,
            s.error_msg = CASE WHEN a.id IS NULL THEN 'DNI no encontrado en alumnos' ELSE NULL END;
        """
    )
    mysql_exec(
        """
        UPDATE staging_excel_regulares s
        INNER JOIN alumnos a ON a.id = s.alumno_id
        SET s.importe_cuota = (
          SELECT ROUND(COALESCE(SUM(ar.importe_referencia), 0), 2)
          FROM alumno_articulo aa
          INNER JOIN articulos ar ON ar.id = aa.articulo_id AND ar.activo = 1
          WHERE aa.alumno_id = a.id
        )
        WHERE s.error_msg IS NULL;
        """
    )
    mysql_exec(
        """
        UPDATE staging_excel_regulares
        SET error_msg = 'Sin artículos asignados o importe 0'
        WHERE error_msg IS NULL AND COALESCE(importe_cuota, 0) <= 0;
        """
    )


def aplicar_import(
    *,
    anio_operativo: int,
    desactivar_no_listados: bool,
    limpiar_cuotas_operativo: bool,
    solo_faltantes: bool,
    dry_run: bool,
) -> None:
    hoy = date.today()
    mes_actual = hoy.month if hoy.year == anio_operativo else 12

    filtro_faltantes = ""
    if solo_faltantes:
        filtro_faltantes = f"""
          AND NOT EXISTS (
            SELECT 1 FROM cuota_mensual cm
            WHERE cm.alumno_id = staging_excel_regulares.alumno_id
              AND cm.anio >= {anio_operativo}
              AND cm.nota LIKE 'Import Excel%'
          )
        """

    if dry_run:
        print("[DRY-RUN] No se escriben cambios en BD.")
    else:

        if desactivar_no_listados and not solo_faltantes:
            mysql_exec(
                f"""
                UPDATE alumnos a
                SET a.activo = 0
                WHERE a.activo = 1
                  AND NOT EXISTS (
                    SELECT 1 FROM staging_excel_regulares s
                    WHERE s.alumno_id = a.id AND s.error_msg IS NULL
                  );
                """
            )
        mysql_exec(
            f"""
            UPDATE alumnos a
            INNER JOIN staging_excel_regulares s ON s.alumno_id = a.id
            SET a.activo = 1,
                a.tipo_alumno = 'regular',
                a.curso = NULLIF(s.carrera, '')
            WHERE s.error_msg IS NULL
            {filtro_faltantes.replace('staging_excel_regulares.alumno_id', 's.alumno_id')};
            """
        )
        mysql_exec(
            """
            UPDATE staging_excel_regulares s
            INNER JOIN alumnos a ON a.id = s.alumno_id
            SET a.observaciones = CONCAT(
              COALESCE(NULLIF(TRIM(a.observaciones), ''), ''),
              CASE WHEN COALESCE(NULLIF(TRIM(a.observaciones), ''), '') = '' THEN '' ELSE ' | ' END,
              'Excel mes abonado: ', s.mes_abonado_raw
            )
            WHERE s.error_msg IS NULL
              AND s.mes_abonado_anio IS NULL
              AND UPPER(COALESCE(s.mes_abonado_raw, '')) IN ('BECA', 'MATRICULA', 'MATRÍCULA');
            """
        )

        if limpiar_cuotas_operativo:
            subq = f"""
                SELECT alumno_id FROM staging_excel_regulares
                WHERE error_msg IS NULL
                {filtro_faltantes}
            """
            mysql_exec(
                f"""
                DELETE pac FROM pago_aplica_cuota pac
                INNER JOIN cuota_mensual cm ON cm.id = pac.cuota_id
                WHERE cm.anio >= {anio_operativo}
                  AND cm.alumno_id IN ({subq});
                DELETE pr FROM pago_registrado pr
                WHERE pr.alumno_id IN ({subq})
                  AND YEAR(pr.fecha_pago) >= {anio_operativo};
                DELETE FROM cuota_mensual
                WHERE anio >= {anio_operativo}
                  AND alumno_id IN ({subq});
                """
            )

    # Generar cuotas 2026 operativo por alumno (en Python para claridad)
    ok_rows = mysql_exec(
        f"""
        SELECT id, alumno_id, importe_cuota, mes_abonado_anio, mes_abonado_mes, mes_abonado_raw, dni
        FROM staging_excel_regulares
        WHERE error_msg IS NULL AND alumno_id IS NOT NULL
        {filtro_faltantes}
        ORDER BY id;
        """,
        fetch=True,
    )
    inserts_cuota = []
    inserts_pago = []
    for line in ok_rows.splitlines():
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) < 7:
            continue
        sid, alumno_id, importe, anio_ab, mes_ab, mes_raw, dni = parts
        alumno_id = int(alumno_id)
        importe = float(importe)
        ultimo_anio = int(anio_ab) if anio_ab and anio_ab != "NULL" else None
        ultimo_mes = int(mes_ab) if mes_ab and mes_ab != "NULL" else None

        for mes in range(1, mes_actual + 1):
            pagada = False
            if ultimo_anio is not None and ultimo_mes is not None:
                if ultimo_anio > anio_operativo or (
                    ultimo_anio == anio_operativo and ultimo_mes >= mes
                ):
                    pagada = True
            estado = "pagada" if pagada else "pendiente"
            saldo = 0.0 if pagada else importe
            fecha_ven = f"{anio_operativo}-{mes:02d}-01"
            nota = f"Import Excel regulares ({mes_raw})"
            inserts_cuota.append(
                (alumno_id, anio_operativo, mes, importe, saldo, estado, nota, pagada, dni)
            )

    for alumno_id, anio, mes, importe, saldo, estado, nota, pagada, dni in inserts_cuota:
        if dry_run:
            continue
        mysql_exec(
            f"""
            INSERT INTO cuota_mensual
              (alumno_id, anio, mes, importe_original, saldo, fecha_vencimiento, estado, nota)
            VALUES
              ({alumno_id}, {anio}, {mes}, {importe:.2f}, {saldo:.2f},
               STR_TO_DATE('{anio}-{mes:02d}-01', '%Y-%m-%d'),
               '{estado}', '{sql_escape(nota)}')
            ON DUPLICATE KEY UPDATE
              importe_original = VALUES(importe_original),
              saldo = VALUES(saldo),
              estado = VALUES(estado),
              nota = VALUES(nota);
            """
        )
        if pagada:
            ref = f"EXCEL:REG:{dni}:{anio}-{mes:02d}"
            mysql_exec(
                f"""
                INSERT INTO pago_registrado
                  (alumno_id, fecha_pago, importe, importe_capital, medio, referencia, nota)
                SELECT {alumno_id},
                       STR_TO_DATE('{anio}-{mes:02d}-01', '%Y-%m-%d'),
                       {importe:.2f}, {importe:.2f}, 'excel', '{sql_escape(ref)}', 'Importado desde Excel regulares'
                WHERE NOT EXISTS (
                  SELECT 1 FROM pago_registrado
                  WHERE alumno_id = {alumno_id} AND referencia = '{sql_escape(ref)}'
                );
                """
            )
            mysql_exec(
                f"""
                INSERT IGNORE INTO pago_aplica_cuota (pago_id, cuota_id, importe_aplicado)
                SELECT pr.id, cm.id, LEAST(pr.importe, cm.importe_original)
                FROM pago_registrado pr
                INNER JOIN cuota_mensual cm
                  ON cm.alumno_id = pr.alumno_id AND cm.anio = {anio} AND cm.mes = {mes}
                WHERE pr.alumno_id = {alumno_id}
                  AND pr.referencia = '{sql_escape(ref)}';
                """
            )

    if not dry_run:
        mysql_exec(
            "UPDATE staging_excel_regulares SET procesado = 1, procesado_en = NOW() WHERE error_msg IS NULL;"
        )


def print_resumen() -> None:
    print("\n=== Resumen staging ===")
    print(
        mysql_exec(
            """
            SELECT CONCAT(
              'Total=', COUNT(*),
              ' | OK=', SUM(error_msg IS NULL),
              ' | Error=', SUM(error_msg IS NOT NULL)
            ) FROM staging_excel_regulares;
            """,
            fetch=True,
        )
    )
    err = mysql_exec(
        "SELECT dni, apellido, nombre, error_msg FROM staging_excel_regulares WHERE error_msg IS NOT NULL LIMIT 20;",
        fetch=True,
    )
    if err:
        print("\nErrores (hasta 20):")
        print(err)


def main() -> None:
    load_env_file()
    parser = argparse.ArgumentParser(description="Importar alumnos regulares desde Excel")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Ruta al .xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Solo staging y resumen, sin aplicar")
    parser.add_argument(
        "--desactivar-no-listados",
        action="store_true",
        help="Pone activo=0 a quienes no están en el Excel",
    )
    parser.add_argument(
        "--limpiar-cuotas-operativo",
        action="store_true",
        help="Borra cuotas/pagos del año operativo antes de importar (solo filas a procesar).",
    )
    parser.add_argument(
        "--solo-faltantes",
        action="store_true",
        help="Solo alumnos que aún no tienen cuotas Import Excel (no toca los 236 ya cargados).",
    )
    parser.add_argument(
        "--corregir-dni-conocidos",
        action="store_true",
        help="Aplica tabla DNI_CORRECCIONES al leer el Excel (typos frecuentes).",
    )
    parser.add_argument(
        "--anio-operativo",
        type=int,
        default=int(os.environ.get("OPERATIVO_ANIO_DESDE", "2026")),
    )
    args = parser.parse_args()

    if not args.excel.is_file():
        print(f"No existe el archivo: {args.excel}")
        sys.exit(1)

    rows = read_excel(args.excel, corregir_dni=args.corregir_dni_conocidos)
    print(f"Filas leídas del Excel: {len(rows)}")
    if args.corregir_dni_conocidos:
        n_corr = sum(1 for r in rows if r.get("dni") != r.get("dni_excel"))
        if n_corr:
            print(f"DNI corregidos automáticamente: {n_corr}")

    ensure_staging_table()
    load_staging(rows, truncate=True)
    match_alumnos()
    print_resumen()

    aplicar_import(
        anio_operativo=args.anio_operativo,
        desactivar_no_listados=args.desactivar_no_listados,
        limpiar_cuotas_operativo=args.limpiar_cuotas_operativo,
        solo_faltantes=args.solo_faltantes,
        dry_run=args.dry_run,
    )

    if args.dry_run:
        print("\nEjecutá sin --dry-run para aplicar. Luego: php tools/recalcular_saldos_cli.php")
    else:
        print("\nImport aplicado. Ejecutá: php tools/recalcular_saldos_cli.php")


if __name__ == "__main__":
    main()
